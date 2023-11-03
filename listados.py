import csv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from variables_globales import USUARIOS, LOCALES, USO_PROMOCIONES
from utils import codigo_local_con_usuario

def listado_promociones_usadas():
    promociones_ordenadas = sorted(USO_PROMOCIONES, key=lambda x: x['cod_cliente'])
    nombre_archivo = "promociones_utilizadas.csv"
    nombre_archivo_excel = "promociones_utilizadas.xlsx"
    campos = ["Nombre", "Codigo Local", "Usuario", "Fecha y Hora", "Codigo Utilizado"]
    if USO_PROMOCIONES == []:
        input("No hay datos para generar un registro. Presione una tecla para volver al menú")
    else: 
        with open(nombre_archivo, 'w', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=campos)
            writer.writeheader()
            for promocion in promociones_ordenadas:
                promocion_modificada = {
                    "Nombre": promocion['nombre'],
                    "Codigo Local": promocion['cod_local'],
                    "Usuario": promocion['cod_cliente'],
                    "Fecha y Hora": promocion['fecha_hora'],
                    "Codigo Utilizado": promocion['codigo_utilizado']
                }
                writer.writerow(promocion_modificada)
        df = pd.read_csv(nombre_archivo)
        df.to_excel(nombre_archivo_excel, index=False)
        
        wb = load_workbook(nombre_archivo_excel)
        ws = wb.active

        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(nombre_archivo_excel)
        print(f"Archivo {nombre_archivo} generado con éxito.")
        print(f"Archivo {nombre_archivo_excel} generado con éxito.")
        
def listado_promociones_local(usuario_encontrado:list):
    codigo_local = codigo_local_con_usuario(usuario_encontrado=usuario_encontrado)
    nombre_archivo = f"promociones_local_{codigo_local}.csv"
    nombre_archivo_excel = f"promociones_local_{codigo_local}.xlsx"
    campos = ["Nombre", "Codigo Local", "Usuario", "Fecha y Hora", "Codigo Utilizado"]
    promociones_local = [promocion for promocion in USO_PROMOCIONES if promocion['cod_local'] == codigo_local]
    if promociones_local == []:
        input("No hay datos para generar un registro. Presione una tecla para volver al menú")
    else:
        with open(nombre_archivo, 'w', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=campos)
            writer.writeheader()
            for promocion in promociones_local:
                promocion_modificada = {
                    "Nombre": promocion['nombre'],
                    "Codigo Local": promocion['cod_local'],
                    "Usuario": promocion['cod_cliente'],
                    "Fecha y Hora": promocion['fecha_hora'],
                    "Codigo Utilizado": promocion['codigo_utilizado']
                }
                writer.writerow(promocion_modificada)
        df = pd.read_csv(nombre_archivo)
        df.to_excel(nombre_archivo_excel, index=False)
        wb = load_workbook(nombre_archivo_excel)
        ws = wb.active

        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(nombre_archivo_excel)
        print(f"Archivo {nombre_archivo} generado con éxito.")
        print(f"Archivo {nombre_archivo_excel} generado con éxito.")

def listado_locales():
    nombre_archivo = "locales.csv"
    nombre_archivo_excel = "locales.xlsx"
    campos = ["Codigo", "Nombre del Local", "Ubicación", "Rubro", "Estado"]
    if LOCALES == []:
        input("No hay datos para generar un registro. Presione una tecla para volver al menú")
    else:
        with open(nombre_archivo, 'w', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=campos)
            writer.writeheader()
            for local in LOCALES:
                local_modificado = {
                    "Codigo": local["cod"],
                    "Nombre del Local": local["nombre"],
                    "Ubicación": local["ubicacion"],
                    "Rubro": local["rubro"],
                    "Estado": 'Activo' if local["estado"] else 'Inactivo'
                }
                writer.writerow(local_modificado)
        df = pd.read_csv(nombre_archivo)
        df.to_excel(nombre_archivo_excel, index=False)
        print(f"Archivo {nombre_archivo} generado con éxito.")
        print(f"Archivo {nombre_archivo_excel} generado con éxito.")
        
def listado_usuarios():
    nombre_archivo = 'usuarios.csv'
    nombre_archivo_excel = 'usuarios.xlsx'
    campos = ['Codigo de usuario', 'Nombre', 'Apellido', 'Correo Electrónico', 'Rol', 'DNI', 'Nro.Teléfono']
    with open(nombre_archivo, 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile,fieldnames=campos)
        writer.writeheader()
        for usuario in USUARIOS:
            usuario_modificado = {
                "Codigo de usuario":usuario['cod'],
                "Nombre": usuario['nombre'],
                "Apellido": usuario['apellido'],
                "Correo Electrónico": usuario['email'],
                "Rol": usuario['rol'],
                "DNI": usuario['dni'],
                "Nro.Teléfono": usuario['telefono']
            }
            writer.writerow(usuario_modificado)
    df = pd.read_csv(nombre_archivo)
    df.to_excel(nombre_archivo_excel, index=False) 
    wb = load_workbook(nombre_archivo_excel)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    wb.save(nombre_archivo_excel)       
    print(f"Archivo {nombre_archivo} generado con éxito. ")
    print(f"Archivo {nombre_archivo_excel} generado con éxito.")
